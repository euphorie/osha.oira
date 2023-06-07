from copy import copy
from euphorie.client import model
from euphorie.client.browser import report
from openpyxl.drawing.image import Image
from openpyxl.styles import Border
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from osha.oira import _
from pkg_resources import resource_filename
from plone import api
from plone.memoize.view import memoize
from plonetheme.nuplone.utils import formatDate
from sqlalchemy import sql
from z3c.saconfig import Session
from zope.i18n import translate

import logging


log = logging.getLogger(__name__)


COLUMN_ORDER = [
    ("risk", "title"),
    ("risk", "priority"),
    ("measure", "action"),
    ("measure", "requirements"),
    ("measure", "planning_start"),
    ("measure", "planning_end"),
    ("measure", "responsible"),
    ("measure", "budget"),
    ("risk", "number"),
    ("module", "title"),
    ("risk", "comment"),
]


class ActionPlanTimeline(report.ActionPlanTimeline):
    title_extra = ""
    combine_keys = ["requirements"]
    extra_cols = [
        (
            "measure",
            "requirements",
            _(
                "label_measure_requirements",
                default="Level of expertise and/or requirements needed",
            ),
        ),
    ]

    columns = [
        ("module", "title", _("label_section", default="Section")),
        (
            "risk",
            "title",
            _("report_timeline_risk_title", default="Description of the risk"),
        ),
        ("risk", "number", _("label_risk_number", default="Risk number")),
        ("risk", "priority", _("report_timeline_priority", default="Priority")),
        (
            "measure",
            "action",
            _(
                "report_timeline_measure",
                default="Measure",
            ),
        ),
        (
            "measure",
            "planning_start",
            _("report_timeline_start_date", default="Start date"),
        ),
        (
            "measure",
            "planning_end",
            _("report_timeline_end_date", default="End date"),
        ),
        (
            "measure",
            "responsible",
            _("report_timeline_responsible", default="Responsible"),
        ),
        ("measure", "budget", _("label_action_plan_budget", default="Budget")),
        (
            None,
            None,
            _(
                "report_timeline_progress",
                default="Status (planned, in process, implemented)",
            ),
        ),
        ("risk", "comment", _("report_timeline_comment", default="Comments")),
    ]

    @property
    @memoize
    def use_solution_description(self):
        country = self.webhelpers.country
        if country in ["it", "fr"]:
            return False
        return True

    def create_workbook(self):
        """Create an Excel workbook containing the all risks and measures."""
        t = lambda txt: translate(txt, context=self.request)  # noqa: E731
        survey = self.webhelpers._survey
        book = Workbook()
        ws1 = book.active
        ws1.title = t(_("report_timeline_title", default="Timeline"))

        header_text = "{title}{extra} - {action_plan}".format(
            title=survey.title,
            extra=self.title_extra.strip(),
            action_plan=t(_("label_action_plan", default="Action Plan")),
        )
        ws1["A1"] = header_text

        font_basic = ws1["A1"].font
        font_large = copy(font_basic)
        font_large.size = 18
        ws1["A1"].font = font_large

        image_filename = resource_filename(
            "osha.oira.client", "resources/oira-logo-colour.png"
        )
        logo = Image(image_filename)
        ws1.add_image(logo, "K1")
        ws1.row_dimensions[1].height = 70
        ws1.merge_cells("A1:K1")

        font_bold = copy(font_basic)
        font_bold.bold = True

        alignment_basic = ws1["A1"].alignment.copy()
        alignment_basic.wrap_text = True
        alignment_basic.vertical = "center"
        alignment_basic.horizontal = "left"
        alignment_header = copy(alignment_basic)
        alignment_header.horizontal = "center"

        ws1.cell(row=1, column=1).alignment = alignment_basic

        b_thin = Side(border_style="thin", color="000000")
        b_double = Side(border_style="medium", color="000000")

        ws1["A2"] = t(_("label_title", default="Title"))
        ws1["A2"].font = font_bold
        ws1["B2"] = self.session.title
        ws1["B2"].fill = PatternFill("solid", fgColor="DDDDDD")
        ws1.merge_cells("B2:C2")
        ws1["E2"] = t(_("label_report_date", default="Date of editing"))
        ws1["E2"].font = font_bold
        ws1["F2"] = formatDate(self.request, self.session.modified)
        ws1["F2"].fill = PatternFill("solid", fgColor="DDDDDD")
        for cell in tuple(ws1.iter_rows(2, 2))[0]:
            cell.alignment = alignment_basic
        ws1.row_dimensions[2].height = 30

        for (column, (type, key, title)) in enumerate(self.columns, 1):
            if key in self.combine_keys:
                continue
            cell = ws1.cell(row=3, column=column)
            cell.value = t(title)
            cell.font = font_bold
            cell.alignment = alignment_header
            # Light baby blue background color
            cell.fill = PatternFill("solid", fgColor="97CDDD")
            cell.border = Border(
                top=b_double, left=b_double, right=b_double, bottom=b_double
            )
            letter = get_column_letter(column)
            if title in ("report_timeline_measure", "report_timeline_risk_title"):
                ws1.column_dimensions[letter].width = len(cell.value) + 50
            elif title in ("label_risk_number",):
                ws1.column_dimensions[letter].width = len(cell.value)
            else:
                ws1.column_dimensions[letter].width = len(cell.value) + 5
        ws1.row_dimensions[3].height = 60

        portal_transforms = api.portal.get_tool("portal_transforms")
        for (row, (module, risk, measure)) in enumerate(self.get_measures(), 4):
            column = 1

            if not getattr(risk, "is_custom_risk", None):
                zodb_node = self.context.restrictedTraverse(risk.zodb_path.split("/"))
            else:
                zodb_node = None

            for (type, key, title) in self.columns + self.extra_cols:
                value = None
                if type == "measure":
                    value = getattr(measure, key, None)
                    if key == "action":
                        if (
                            self.use_solution_description
                            and zodb_node
                            and measure.solution_id in zodb_node
                        ):
                            description = zodb_node[measure.solution_id].description
                            value = f"{description}\n{value}"
                        value = portal_transforms.convertToData("text/plain", value)
                elif type == "risk":
                    value = getattr(risk, key, None)
                    if key == "priority":
                        value = self.priority_name(value)
                    elif key == "title":
                        if zodb_node is None:
                            value = getattr(risk, key, None)
                        elif (
                            zodb_node.problem_description
                            and zodb_node.problem_description.strip()
                        ):
                            value = zodb_node.problem_description
                    elif key == "number":
                        if risk.is_custom_risk:
                            num_elems = value.split(".")
                            value = ".".join(["Ω"] + num_elems[1:])

                elif type == "module":
                    if key == "title" and module.depth > 1:
                        titles = []
                        m = module
                        while m:
                            title = getattr(m, "title", None)
                            if title:
                                titles.append(m.title)
                            m = m.parent
                        titles.reverse()
                        value = ", ".join(titles)
                    else:
                        if module.zodb_path == "custom-risks":
                            lang = getattr(self.request, "LANGUAGE", "en")
                            if "-" in lang:
                                elems = lang.split("-")
                                lang = f"{elems[0]}_{elems[1].upper()}"
                            value = translate(
                                _("Custom risks", default="Custom risks"),
                                target_language=lang,
                            )
                        else:
                            value = getattr(module, key, None)
                if key in self.combine_keys:
                    if value is not None:
                        # osha wants to combine action_plan (col 5 / E),
                        # and requirements in one cell
                        if not ws1.cell(row=row, column=5).value:
                            ws1.cell(row=row, column=5).value = ""
                        ws1.cell(row=row, column=5).value += "\r\n" + value
                else:
                    cell = ws1.cell(row=row, column=column)
                    if value is not None:
                        if key == "number":
                            # force string
                            cell.set_explicit_value(value)
                        else:
                            cell.value = value
                        if key == "budget":
                            cell.style = "Comma"
                    cell.alignment = alignment_basic
                    cell.border = Border(
                        top=b_thin, left=b_thin, right=b_thin, bottom=b_thin
                    )

                    column += 1
        ws1.freeze_panes = "A4"
        ws1.set_printer_settings(paper_size=ws1.PAPERSIZE_A4, orientation="landscape")
        return book

    def get_measures(self):
        """Find all data that should be included in the report.

        The data is returned as a list of tuples containing a
        :py:class:`Module <euphorie.client.model.Module>`,
        :py:class:`Risk <euphorie.client.model.Risk>` and
        :py:class:`ActionPlan <euphorie.client.model.ActionPlan>`. Each
        entry in the list will correspond to a row in the generated Excel
        file.

        This implementation differs from Euphorie in its ordering:
        it sorts on risk priority instead of start date.
        """
        query = (
            Session.query(model.Module, model.Risk, model.ActionPlan)
            .select_from(model.Module)
            .filter(
                sql.and_(
                    model.Module.session == self.session,
                    model.Module.profile_index > -1,
                )
            )
            .filter(sql.not_(model.SKIPPED_PARENTS))
            .filter(
                sql.or_(
                    model.MODULE_WITH_RISK_OR_TOP5_FILTER,
                    model.RISK_PRESENT_OR_TOP5_FILTER,
                )
            )
            .join(model.Risk, model.Risk.parent_id == model.Module.id)
            .join(
                (
                    model.ActionPlan,
                    sql.and_(
                        model.ActionPlan.risk_id == model.Risk.id,
                        sql.or_(
                            model.ActionPlan.plan_type == "measure_standard",
                            model.ActionPlan.plan_type == "measure_custom",
                        ),
                    ),
                )
            )
            .order_by(
                sql.case(
                    value=model.Risk.priority, whens={"high": 0, "medium": 1}, else_=2
                ),
                model.Risk.path,
            )
        )
        return [
            t
            for t in query.all()
            if (
                (
                    t[-1].planning_start is not None
                    or t[-1].planning_end is not None
                    or t[-1].responsible is not None
                    or t[-1].requirements is not None
                    or t[-1].budget is not None
                    or t[-1].action is not None
                )
                and (t[1].identification == "no" or t[1].risk_type == "top5")
            )
        ]


def node_title(node, zodbnode):
    # 2885: Non-present risks and unanswered risks are shown affirmatively,
    # i.e 'title'
    if node.type != "risk" or node.identification in ["n/a", "yes", None]:
        return node.title
    # The other two groups of risks are shown negatively, i.e
    # 'problem_description'
    if zodbnode.problem_description and zodbnode.problem_description.strip():
        return zodbnode.problem_description
    return node.title
