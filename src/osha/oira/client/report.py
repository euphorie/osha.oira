# coding=utf-8
from euphorie.client import model
from euphorie.client import report
from five import grok
from openpyxl.cell import get_column_letter
from openpyxl.workbook import Workbook
from osha.oira import _
from osha.oira.client.interfaces import IOSHAClientSkinLayer
from sqlalchemy import sql
from z3c.saconfig import Session
from zope.i18n import translate
import logging


log = logging.getLogger(__name__)

grok.templatedir("templates")


COLUMN_ORDER = [
    ("risk", "title"),
    ("risk", "priority"),
    ("measure", "action_plan"),
    ("measure", "prevention_plan"),
    ("measure", "requirements"),
    ("measure", "planning_start"),
    ("measure", "planning_end"),
    ("measure", "responsible"),
    ("measure", "budget"),
    ("risk", "number"),
    ("module", "title"),
    ("risk", "comment"),
]


class ActionPlanTimeline(report.ActionPlanTimeline):
    grok.layer(IOSHAClientSkinLayer)

    combine_keys = ["prevention_plan", "requirements"]
    columns = sorted(
        (
            col
            for col in report.ActionPlanTimeline.columns
            if (col[0], col[1]) in COLUMN_ORDER
        ),
        key=lambda d, co=COLUMN_ORDER: co.index((d[0], d[1])),
    )
    extra_cols = [x for x in columns if x[1] in combine_keys]
    columns = [x for x in columns if x[1] not in combine_keys]
    columns[2] = (
        "measure",
        "action_plan",
        _("report_timeline_measure", default=u"Measure"),
    )
    columns[3] = (
        "measure",
        "planning_start",
        _("report_timeline_start_date", default=u"Start date"),
    )
    columns[4] = (
        "measure",
        "planning_end",
        _("report_timeline_end_date", default=u"End date"),
    )
    columns[5] = (
        "measure",
        "responsible",
        _("report_timeline_responsible", default=u"Responsible"),
    )
    columns.insert(
        -1,
        (
            None,
            None,
            _(
                "report_timeline_progress",
                default=u"Status (planned, in process, implemented)",
            ),
        ),
    )

    def create_workbook(self):
        """Create an Excel workbook containing the all risks and measures.
        """
        t = lambda txt: translate(txt, context=self.request)
        book = Workbook()
        sheet = book.worksheets[0]
        sheet.title = t(_("report_timeline_title", default=u"Timeline"))
        sheet.default_column_dimension.auto_size = True

        for (column, (type, key, title)) in enumerate(self.columns):
            if key in self.combine_keys:
                continue
            cell = sheet.cell(row=0, column=column)
            cell.value = t(title)
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = True
            letter = get_column_letter(column + 1)
            if title == "report_timeline_measure":
                sheet.column_dimensions[letter].width = len(cell.value) + 50
            else:
                sheet.column_dimensions[letter].width = len(cell.value) + 5

        for (row, (module, risk, measure)) in enumerate(self.get_measures(), 1):
            column = 0

            if not getattr(risk, "is_custom_risk", None):
                zodb_node = self.context.restrictedTraverse(
                    risk.zodb_path.split("/")
                )
            else:
                zodb_node = None

            for (type, key, title) in self.columns + self.extra_cols:
                value = None
                if type == "measure":
                    value = getattr(measure, key, None)
                elif type == "risk":
                    value = getattr(risk, key, None)
                    if key == "priority":
                        value = self.priority_name(value)
                    elif key == "title":
                        if zodb_node is None:
                            value = getattr(risk, key, None)
                        elif (
                            zodb_node.problem_description
                            and zodb_node.problem_description.strip()
                        ):
                            value = zodb_node.problem_description
                    elif key == "number":
                        if risk.is_custom_risk:
                            num_elems = value.split(".")
                            value = u".".join([u"Ω"] + num_elems[1:])

                elif type == "module":
                    if key == "title" and module.depth > 1:
                        titles = []
                        m = module
                        while m:
                            title = getattr(m, "title", None)
                            if title:
                                titles.append(m.title)
                            m = m.parent
                        titles.reverse()
                        value = ", ".join(titles)
                    else:
                        if module.zodb_path == "custom-risks":
                            lang = getattr(self.request, "LANGUAGE", "en")
                            if "-" in lang:
                                elems = lang.split("-")
                                lang = "{0}_{1}".format(elems[0], elems[1].upper())
                            value = translate(
                                _("title_other_risks", default=u"Added risks (by you)"),
                                target_language=lang,
                            )
                        else:
                            value = getattr(module, key, None)

                sheet.cell(
                    row=row, column=column
                ).style.alignment.wrap_text = True  # style
                if key in self.combine_keys and value is not None:
                    # osha wants to combine action_plan (col 3),
                    # prevention_plan and requirements in one cell
                    if not sheet.cell(row=row, column=2).value:
                        sheet.cell(row=row, column=2).value = u""
                    sheet.cell(row=row, column=2).value += "\r\n" + value
                    continue

                if value is not None:
                    sheet.cell(row=row, column=column).value = value
                column += 1
        return book

    def get_measures(self):
        """Find all data that should be included in the report.

        The data is returned as a list of tuples containing a
        :py:class:`Module <euphorie.client.model.Module>`,
        :py:class:`Risk <euphorie.client.model.Risk>` and
        :py:class:`ActionPlan <euphorie.client.model.ActionPlan>`. Each
        entry in the list will correspond to a row in the generated Excel
        file.

        This implementation differs from Euphorie in its ordering:
        it sorts on risk priority instead of start date.
        """
        query = (
            Session.query(model.Module, model.Risk, model.ActionPlan)
            .filter(
                sql.and_(
                    model.Module.session == self.session,
                    model.Module.profile_index > -1,
                )
            )
            .filter(sql.not_(model.SKIPPED_PARENTS))
            .filter(
                sql.or_(
                    model.MODULE_WITH_RISK_OR_TOP5_FILTER,
                    model.RISK_PRESENT_OR_TOP5_FILTER,
                )
            )
            .join(
                (
                    model.Risk,
                    sql.and_(
                        model.Risk.path.startswith(model.Module.path),
                        model.Risk.depth == model.Module.depth + 1,
                        model.Risk.session == self.session,
                    ),
                )
            )
            .join((model.ActionPlan, model.ActionPlan.risk_id == model.Risk.id))
            .order_by(
                sql.case(
                    value=model.Risk.priority, whens={"high": 0, "medium": 1}, else_=2
                ),
                model.Risk.path,
            )
        )
        return [
            t
            for t in query.all()
            if (
                (
                    t[-1].planning_start is not None
                    or t[-1].planning_end is not None
                    or t[-1].responsible is not None
                    or t[-1].prevention_plan is not None
                    or t[-1].requirements is not None
                    or t[-1].budget is not None
                    or t[-1].action_plan is not None
                )
                and (t[1].identification == "no" or t[1].risk_type == "top5")
            )
        ]


def node_title(node, zodbnode):
    # 2885: Non-present risks and unanswered risks are shown affirmatively,
    # i.e 'title'
    if node.type != "risk" or node.identification in [u"n/a", u"yes", None]:
        return node.title
    # The other two groups of risks are shown negatively, i.e
    # 'problem_description'
    if zodbnode.problem_description and zodbnode.problem_description.strip():
        return zodbnode.problem_description
    return node.title
